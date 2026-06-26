"""Writes the price backfill to an .xlsx workbook, one row per (date, stock).

The sheet mirrors the user's template: a two-row grouped header
(Future / Option (CE) / Option (PE) / NET DAILY P&L / VIX bands over OHLC
sub-headers), divider columns, and placeholder columns (prices, P&L, IV) the
user fills in later — the bot leaves those blank.

Saves after every stock so an interrupted/rate-limited run resumes cleanly: on
restart we skip any (DATE, SCRIPT) already present. COLUMNS is the single source
of truth for column order, the two header rows, and which row-dict key (if any)
feeds each column.
"""

from __future__ import annotations

import os

from openpyxl import Workbook, load_workbook

from config import Config

# (group header [row 1], sub-header [row 2], row-dict key or None) per column, in order.
# key=None -> a divider or a placeholder column the bot leaves blank.
COLUMNS: list[tuple[str, str, str | None]] = [
    ("Date",   "ALL",    "DATE"),
    ("Script", "Equity", "SCRIPT"),
    ("",       "",       None),                 # divider
    ("",       "QTY",    "QTY"),
    ("",       "",       None),                 # divider
    ("Future", "BUYING Price",    None),
    ("Future", "OPEN",            "FUT_OPEN"),
    ("Future", "High",            "FUT_HIGH"),
    ("Future", "Low",             "FUT_LOW"),
    ("Future", "Close",           "FUT_CLOSE"),
    ("Future", "DAILY P&L OPEN",  None),
    ("Future", "DAILY P&L High",  None),
    ("Future", "DAILY P&L Low",   None),
    ("Future", "DAILY P&L Close", None),
    ("",       "",       None),                 # divider
    ("Option (CE)", "Strike",          "CE_STRIKE"),
    ("Option (CE)", "Selling Price",   None),
    ("Option (CE)", "OPEN",            "CE_OPEN"),
    ("Option (CE)", "High",            "CE_HIGH"),
    ("Option (CE)", "Low",             "CE_LOW"),
    ("Option (CE)", "Close",           "CE_CLOSE"),
    ("Option (CE)", "DAILY P&L OPEN",  None),
    ("Option (CE)", "DAILY P&L High",  None),
    ("Option (CE)", "DAILY P&L Low",   None),
    ("Option (CE)", "DAILY P&L Close", None),
    ("Option (CE)", "Open IV",         None),
    ("Option (CE)", "High IV",         None),
    ("Option (CE)", "Low IV",          None),
    ("Option (CE)", "Close IV",        None),
    ("",       "",       None),                 # divider
    ("Option (PE)", "Strike",          "PE_STRIKE"),
    ("Option (PE)", "Buying Price",    None),
    ("Option (PE)", "OPEN",            "PE_OPEN"),
    ("Option (PE)", "High",            "PE_HIGH"),
    ("Option (PE)", "Low",             "PE_LOW"),
    ("Option (PE)", "Close",           "PE_CLOSE"),
    ("Option (PE)", "DAILY P&L OPEN",  None),
    ("Option (PE)", "DAILY P&L High",  None),
    ("Option (PE)", "DAILY P&L Low",   None),
    ("Option (PE)", "DAILY P&L Close", None),
    ("Option (PE)", "Open IV",         None),
    ("Option (PE)", "High IV",         None),
    ("Option (PE)", "Low IV",          None),
    ("Option (PE)", "Close IV",        None),
    ("",       "",       None),                 # divider
    ("NET DAILY P&L", "Daily P&L Open",  None),
    ("NET DAILY P&L", "Daily P&L High",  None),
    ("NET DAILY P&L", "Daily P&L Low",   None),
    ("NET DAILY P&L", "Daily P&L Close", None),
    ("",       "",       None),                 # divider
    ("VIX", "Open",  "VIX_OPEN"),
    ("VIX", "High",  "VIX_HIGH"),
    ("VIX", "Low",   "VIX_LOW"),
    ("VIX", "Close", "VIX_CLOSE"),
]

KEYS = [c[2] for c in COLUMNS]          # column-ordered keys (None = blank column)
HEADER_ROWS = 2
_DATE_COL = KEYS.index("DATE")
_SCRIPT_COL = KEYS.index("SCRIPT")


class ExcelWriter:
    def __init__(self, path: str = Config.OUTPUT_XLSX, append: bool = True) -> None:
        """append=True loads/extends an existing file (resume-aware);
        append=False starts a fresh workbook, overwriting the path on save."""
        self.path = path
        if append and os.path.exists(path):
            self.wb = load_workbook(path)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = Config.SHEET_NAME
            self._write_header()
        self._keys = self._scan_existing_keys()

    # -- header ------------------------------------------------------------
    def _write_header(self) -> None:
        self.ws.append([c[0] for c in COLUMNS])   # row 1: section bands
        self.ws.append([c[1] for c in COLUMNS])   # row 2: sub-headers
        self._merge_bands()

    def _merge_bands(self) -> None:
        """Merge each run of adjacent columns sharing the same non-empty group label."""
        n = len(COLUMNS)
        start = 0
        for i in range(1, n + 1):
            if i < n and COLUMNS[i][0] == COLUMNS[start][0]:
                continue
            group = COLUMNS[start][0]
            if group and i - start > 1:
                self.ws.merge_cells(
                    start_row=1, start_column=start + 1, end_row=1, end_column=i
                )
            start = i

    # -- resume ------------------------------------------------------------
    def _scan_existing_keys(self) -> set[tuple[str, str]]:
        keys: set[tuple[str, str]] = set()
        for row in self.ws.iter_rows(min_row=HEADER_ROWS + 1, values_only=True):
            if row and row[_DATE_COL] and row[_SCRIPT_COL]:
                keys.add((str(row[_DATE_COL]), str(row[_SCRIPT_COL])))
        return keys

    def has(self, date_str: str, script: str) -> bool:
        return (date_str, script) in self._keys

    def add_rows(self, rows: list[dict]) -> None:
        for r in rows:
            self.ws.append([r.get(k) if k else None for k in KEYS])
            if r.get("DATE") and r.get("SCRIPT"):
                self._keys.add((str(r["DATE"]), str(r["SCRIPT"])))

    def save(self) -> None:
        self.wb.save(self.path)
