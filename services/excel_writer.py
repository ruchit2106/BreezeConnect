"""Writes results to an .xlsx workbook, one stock = two rows (CALL then PUT).

Saves after every stock so an interrupted/rate-limited run resumes cleanly:
on restart we skip any (DATE, SCRIPT) already present.

KEYS are the internal row-dict keys; the header row shown in Excel is built from
them with the snapshot time appended (e.g. "OPEN_IV (at 9:30 AM)"), derived from
Config.OPEN_TIME / CLOSE_TIME so the labels always match what was actually sampled.
"""

from __future__ import annotations

import os
from datetime import datetime

from openpyxl import Workbook, load_workbook

from config import Config

# Internal keys used to pull values out of each row dict (order = column order).
KEYS = [
    "DATE", "DAY", "SCRIPT", "RIGHT", "EXPIRY",
    "SPOT_OPEN", "OPEN_STRIKE", "OPEN_PREMIUM", "OPEN_IV", "OPEN_DELTA",
    "SPOT_CLOSE", "CLOSE_STRIKE", "CLOSE_PREMIUM", "CLOSE_IV", "CLOSE_DELTA",
    "VIX_OPEN", "VIX_CLOSE",
]


def _fmt_time(hms: str) -> str:
    """'09:30:00' -> '9:30 AM',  '15:15:00' -> '3:15 PM'."""
    return datetime.strptime(hms, "%H:%M:%S").strftime("%I:%M %p").lstrip("0")


def _display_headers() -> list[str]:
    """Human-readable header row with the snapshot time on each timed column."""
    o, c = _fmt_time(Config.OPEN_TIME), _fmt_time(Config.CLOSE_TIME)
    return [
        "DATE", "DAY", "SCRIPT", "RIGHT", "EXPIRY",
        f"SPOT_OPEN (at {o})",
        f"OPEN_STRIKE (at {o})", f"OPEN_PREMIUM (at {o})", f"OPEN_IV (at {o})", f"OPEN_DELTA (at {o})",
        f"SPOT_CLOSE (at {c})",
        f"CLOSE_STRIKE (at {c})", f"CLOSE_PREMIUM (at {c})", f"CLOSE_IV (at {c})", f"CLOSE_DELTA (at {c})",
        f"VIX_OPEN (at {o})", f"VIX_CLOSE (at {c})",
    ]


class ExcelWriter:
    def __init__(self, path: str = Config.OUTPUT_XLSX, append: bool = True) -> None:
        """append=True loads/extends an existing file (resume-aware);
        append=False starts a fresh workbook, overwriting the path on save."""
        self.path = path
        if append and os.path.exists(path):
            self.wb = load_workbook(path)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = Config.SHEET_NAME
            self.ws.append(_display_headers())
        self._keys = self._scan_existing_keys()

    def _scan_existing_keys(self) -> set[tuple[str, str]]:
        keys: set[tuple[str, str]] = set()
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] and row[2]:
                keys.add((str(row[0]), str(row[2])))
        return keys

    def has(self, date_str: str, script: str) -> bool:
        return (date_str, script) in self._keys

    def add_rows(self, rows: list[dict]) -> None:
        for r in rows:
            self.ws.append([r.get(k) for k in KEYS])
            if r.get("DATE") and r.get("SCRIPT"):
                self._keys.add((str(r["DATE"]), str(r["SCRIPT"])))

    def save(self) -> None:
        self.wb.save(self.path)
